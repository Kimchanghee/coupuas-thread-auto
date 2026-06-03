# -*- coding: utf-8 -*-
"""Build threads_coupang_aggro_analysis.xlsx from collected Threads samples."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "threads_coupang_aggro_analysis.xlsx")

# ---- styling helpers ----
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
ALT_FILL = PatternFill("solid", fgColor="F2F5FB")

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = CENTER
        cell.border = BORDER

def finalize(ws, header_row, ncols, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.sheet_view.showGridLines = False

wb = Workbook()

# =========================================================
# Sheet 1: 표본 (Samples)
# =========================================================
ws = wb.active
ws.title = "표본"
ws["A1"] = "스레드 쿠팡파트너스 어그로 글 표본 (1차 본문 + 2차 댓글)"
ws["A1"].font = TITLE_FONT
ws["A2"] = "조사일 2026-06-03 · 검색어 '쿠팡파트너스'/'쿠파스' · 캠페인 26행(개별 1차·2차 글 35+개)"
ws["A2"].font = Font(italic=True, color="595959", size=10)

headers = ["번호", "작성자", "시점", "상품/주제", "1차 후킹유형", "1차 본문 요약(어그로)",
           "어그로강도", "2차 배치방식", "2차 핵심문구", "링크형태", "지표 ♥/💬/🔁/↗", "비고"]
HR = 4
for c, h in enumerate(headers, 1):
    ws.cell(row=HR, column=c, value=h)
style_header(ws, HR, len(headers))

rows = [
 [1,"_tsxiii","2026-01-16","코칭·오픈채팅(수익인증)","A 수익인증","\"급여날 거의 3천 채웠어\" + 입금 27,644,560원 스샷","★★★","④ 톡방/첫댓글","\"톡방 주소는 댓글에 남겨놨어!\"","오픈채팅","377/103/33/30","최고 인게이지먼트. 상품 아닌 코칭 퍼널"],
 [2,"_tsxiii","2025-11-27","무료코칭(수익인증)","A 수익인증","\"정산오류로 이틀치 한번에\" + 무료코칭 예고","★★","④ DM/댓글","\"한사람 한사람 무료코칭 해줄께\"","DM","105/80/28/7","친근·겸손 톤으로 신뢰 구축"],
 [3,"pickmeup__shop","2026-01-12","유튜브 노하우(일지)","A 수익인증","\"쿠파스기록 46일차\" + 수익 ₩3,190,375 리포트","★★","유튜브/프로필 유도","\"유튜브에 리얼 노하우 남기는 중\"","유튜브","61/40/15/2","합산금액 ₩109,339,898 노출"],
 [4,"pickmeup__shop","2026-01-08","노하우(일지)","E 실용/F 모집","\"쿠파스 감 잡은 계기: 트래픽을 많이 만든다\"","★","유튜브 유도","-","-","43/28/6/7","교육형 신뢰 축적"],
 [5,"hae_hyo_","2026-04-13","자동화 솔루션(모집)","A 수익인증/F 모집","\"하루 30분 세팅, 자동화 관심있다면 스하리\"","★★","④ DM + 베이트","\"스하리 해주세요^^\"","DM","213/267/109/3","리포트 스샷 첨부"],
 [6,"growth.jin","2026-04-30","코칭(모집)","F 모집","\"저와 함께면 성공, 33만 팔로워 인사이트, 따라오세요\"","★★","④ DM/팔로우","\"맨땅에 헤딩한 인사이트 나눔\"","-","116/76/25/5","권위(팔로워수) 어필"],
 [7,"introverted_digital_nomad","2026-01-30","노하우 콘텐츠(정보)","E 실용","\"쇼핑쇼츠로 월 300 버는 11단계\"","★★","정보형(신뢰 구축)","-","-","300/25/125/137","리포스트 137 = 확산 강함"],
 [8,"introverted_digital_nomad","2026-05-15","자동화 운영일지","(일지)","\"자동화 2차 사이클 9일차, 계정 6개 테스트\"","★","-","-","-","20/15/3/1","운영 투명 공개형"],
 [9,"unnee___","2025-09-12","영상 상품","B 호기심","\"진짜 신박한거 소개할게!\" (제품 비공개)","★★★","① 연결글 고지","\"본 영상은 쿠팡파트너스 활동의 일환…\"","영상/프로필","457/17/43/213","좋아요 457 고후킹"],
 [10,"army_97554","2025-06-21","생활용품(변기세정)","D 긴급","\"⚠️긴.급.속.보⚠️ 가격 고장났음, 꼭 2개 선택\"","★★★","① 본문/연결 링크+베이트","\"스+하+릿 = 초고속 2배반사\"","link.coupang/a/cAr","19/9/11/4","가격조작 뉘앙스 + 옵션강매"],
 [11,"da_mood.zip","21시간","흑임자 스무디(비밀재료)","B 호기심","\"요즘 다들 흑임자 먹는 이유가 있었네;;\"","★★","① 연결글 레시피+링크","\"맛·식감 살리는 비밀재료 👇\"","link.coupang/a/egN","2/1","레시피로 위장, 핵심재료=상품"],
 [12,"da_mood.zip","21시간","넥크림(목주름)","B 호기심","\"44살 정유미 동안비결, 10년째 넥크림\"","★★","① 연결글 링크","\"목까지 바르는 중 👇\"","link.coupang/a/efp","2/1","연예인 떡밥 후킹"],
 [13,"orion_on82","1시간","자라 향수","C 스토리·유머","\"뒤태 여신의 치명적인 유혹…\" (반전 개그)","★★","① 연결글 링크","\"진짜 전세계 자라 1위 향수는 맞음 👇\"","link.coupang/a/egJ","-","완결 서사, 제품이 펀치라인"],
 [14,"orion_on82","5분","푸마 스니커즈","D 긴급·희소","\"26년 신상 대란, 리셀가 붙기전 막차타라\"","★★","① 연결글 링크","\"가격도 이정도면 끝난거지\"","link.coupang/a/egQ","1","FOMO·트렌드(발레코어)"],
 [15,"weondaegyu82","1일","골반 운동기구","E 실용","1차 \"좀비요가, 기묘한 자세\" → 2차 \"골반 5~10분 몸짱\"","★★","① 연결글 (광고)+링크","\"예쁜이 엉더미 필수템\"","link.coupang/a/egP","176/14/23/233","리포스트 233 확산 큼"],
 [16,"ggultem.memo","56분","선글라스 파우치","C 스토리","\"일본여행 가방마다 이게…난리난 파우치\"","★★","① 연결글 추천템 링크","\"영상에 나온 추천템 보러가기\"","link.coupang/a/egN","1/3/2","'광고' 라벨 표기형"],
 [17,"woowoo_1106","46분","추천템 묶음","E/기타","고지문구로 시작 + 추천템 모음","★","① 연결글 멀티링크","\"영상에 나온 추천템 보러가기\" ×2","link.coupang/a/egN ×2","1/1","링크 2개 병렬 배치"],
 [18,"richrich7821","3분","손선풍기(계절가전)","D 긴급(폭염)","\"역대급 폭염 이거 없으면 후회…한정수량 품절주의\"","★★★","③ 고정댓글/프로필","\"최저가 링크 [고정 댓글]/[프로필 링크]\"","고정댓글/프로필","-","제목/본문 분리, 해시태그 다수"],
 [19,"stay___lab","34분","숙소(여행)","E 실용","\"국내에서 제일 긴 야외 수영장 숙소\"","★★","① 연결글 (광고)+링크","\"자세한 정보는 링크 통해서 확인\"","link.coupang/a/efB","8/3/1","여행후기→숙소 제휴"],
 [20,"kkultemi","12분","락피쉬 신발","C 공감스토리","\"매장서 충동구매함;; 양말 사려다 신어보고 끝남\"","★★","① 연결글 링크 ×2","\"신고 나오면 거울 5번 봄 👇\"","link.coupang/a/egM ×2","1","공감형 충동구매 서사"],
 [21,"bebae_picks","20분","쯔유/우동 재료","E 실용","\"배달앱 켜기 전에 이거 먼저 봐요\"","★★","① 연결글 레시피+링크+상품카드","\"내가 사용한 반칙 재료 ⬇️\"","link.coupang/a/egP","1/1","쿠팡 상품카드(혼쯔유) 첨부"],
 [22,"leo_smart_setuplife","33분","수영모(유머)","C 스토리·유머","\"내 아이 어디있는지 바로 알아 / 개웃기네\"","★★","① 연결글 링크 ×2 + 이모지도배","\"눈에 파바박…개웃겨 😍×16\"","link.coupang/a/egP ×2","3/1/1","이모지 도배형 유도"],
 [23,"yks1220__","45분","공중부양 하우스(인테리어)","C 감성","\"보고있으면 동화에 빠진듯 힐링됨\"","★","① 연결글 링크+상품카드","\"화날때마다 보는데 힐링됨\"","link.coupang/a/egO","2/1","'광고' 라벨 + 감성 소구"],
 [24,"diagmy_","19시간","마늘쫑(식재료)","E 친근정보","\"햇 마늘쫑 철, 싼거보다 좋은걸로 사~\"","★","① 본문말미 링크","\"쿠팡파트너스 수수료를 받아요\"","link.coupang/a/eft","1","구어체 친근 톤"],
 [25,"ggdldmswjd25","18~27분","정보영상(한국사 만화 등) 001~003","B/정보","\"상품 확인하려면 프로필링크+번호 검색\"","★","② 프로필링크+번호","\"프로필링크 클릭 후 \\\"003번\\\" 검색\"","프로필 링크트리","-","번호 시리즈로 다상품 관리"],
 [26,"vellichor.pogeun","21시간","수세미/주방용품","(영상)","고지문구 + \"제품 확인하려면 프로필 링크 클릭\"","★","② 프로필링크","\"프로필 링크를 클릭해주세요🙂\"","프로필 링크","-","해시태그(#수세미 등)"],
]

r = HR + 1
for row in rows:
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = BORDER
        if c in (1, 5, 7, 8, 10, 11):
            cell.alignment = CENTER
        else:
            cell.alignment = WRAP
    if (r - HR) % 2 == 0:
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).fill = ALT_FILL
    r += 1

finalize(ws, HR, len(headers), [5, 20, 11, 20, 14, 40, 9, 18, 34, 20, 13, 26])

# =========================================================
# Sheet 2: 1차 후킹 유형
# =========================================================
ws2 = wb.create_sheet("1차_후킹유형")
ws2["A1"] = "1차 본문 어그로(후킹) 유형 6종"
ws2["A1"].font = TITLE_FONT
h2 = ["유형", "정의", "후킹 트리거", "강도", "대표 예시", "이어지는 2차"]
HR2 = 3
for c, h in enumerate(h2, 1):
    ws2.cell(row=HR2, column=c, value=h)
style_header(ws2, HR2, len(h2))
data2 = [
 ["A 수익인증","구체 금액·입금/리포트 스샷으로 결과 과시","부러움·욕망(나도 저렇게)","★★★","_tsxiii 입금 2,766만원 / pickmeup 수익 319만원","톡방·DM·코칭(④)"],
 ["B 호기심·티저","제품명·링크 숨기고 궁금증만 남김","정보격차(뭔데?)","★★★","unnee \"신박한거 소개\" / da_mood \"흑임자 이유\"","연결글 링크(①)"],
 ["C 스토리·유머","완결 서사/반전 개그에 제품을 주인공으로","재미·공감, 광고 망각","★★","orion \"뒤태 여신…\" / kkultemi 충동구매","연결글 링크(①)"],
 ["D 긴급·희소","품절·대란·가격고장으로 즉시구매 압박","FOMO(지금 아니면 손해)","★★★","army \"긴급속보 가격고장\" / richrich \"한정수량\"","연결글/고정댓글(①③)"],
 ["E 실용·정보","레시피·운동·여행 정보의 마지막 재료/도구가 상품","저장각·유용함","★★","bebae 우동레시피 / weondaegyu 골반운동","연결글 링크(①)"],
 ["F 멘토·모집","권위+무료 미끼로 코칭/유료방 모집","권위·소속감","★★","growth.jin \"33만 팔로워, 따라오세요\"","DM·팔로우(④)"],
]
r = HR2 + 1
for row in data2:
    for c, val in enumerate(row, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.border = BORDER
        cell.alignment = CENTER if c in (1, 4) else WRAP
    if (r - HR2) % 2 == 0:
        for c in range(1, len(h2) + 1):
            ws2.cell(row=r, column=c).fill = ALT_FILL
    r += 1
finalize(ws2, HR2, len(h2), [16, 34, 24, 8, 38, 22])

# =========================================================
# Sheet 3: 2차 배치방식
# =========================================================
ws3 = wb.create_sheet("2차_배치방식")
ws3["A1"] = "2차(댓글/연결글) 상품·링크 배치 메커니즘 4종"
ws3["A1"].font = TITLE_FONT
h3 = ["메커니즘", "설명", "표준 골격", "대표 예시", "비고"]
HR3 = 3
for c, h in enumerate(h3, 1):
    ws3.cell(row=HR3, column=c, value=h)
style_header(ws3, HR3, len(h3))
data3 = [
 ["① 연결 스레드(자기답글) 직접링크","1차 바로 아래 자기 답글 \"2/2\"에 쿠팡 링크","[고지문구]+[카피+👇]+[link.coupang]+[상품카드]+[가격중립]","bebae, da_mood, orion, kkultemi, leo, yks","가장 흔함. 링크 1~2개"],
 ["② 프로필 링크 + 번호 검색","본문/댓글에 링크 안 넣고 프로필 링크트리로 우회","\"프로필링크 클릭 후 0XX번 검색\"","ggdldmswjd25(001~003), vellichor","외부링크 페널티 회피, 다상품 관리"],
 ["③ 고정 댓글(핀)","구매 링크를 고정 댓글에 배치","\"최저가 링크는 [고정 댓글]에서 확인\"","richrich7821","프로필 링크와 병행 안내"],
 ["④ 톡방·DM 유도","상품이 아닌 오픈채팅/코칭으로 퍼널","\"톡방 주소는 댓글에\" / \"스하리 해주세요\"","_tsxiii, hae_hyo_, growth.jin","수익인증·모집형 전용"],
]
r = HR3 + 1
for row in data3:
    for c, val in enumerate(row, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        cell.border = BORDER
        cell.alignment = WRAP
    if (r - HR3) % 2 == 0:
        for c in range(1, len(h3) + 1):
            ws3.cell(row=r, column=c).fill = ALT_FILL
    r += 1
finalize(ws3, HR3, len(h3), [26, 34, 40, 28, 26])

# 2차 공통 구성요소
base = r + 1
ws3.cell(row=base, column=1, value="2차 공통 구성요소 (체크리스트)").font = Font(bold=True, size=12, color="1F3864")
hh = ["요소", "역할", "실제 문구 예"]
hr = base + 1
for c, h in enumerate(hh, 1):
    ws3.cell(row=hr, column=c, value=h)
style_header(ws3, hr, len(hh))
comp = [
 ["의무 고지문구","공정위 법적 표기","\"이 포스팅/영상은 쿠팡파트너스 활동의 일환으로 일정액의 수수료를 제공받습니다\""],
 ["가격중립 안심문구","구매 저항 완화","\"구매자 가격 변동 없음\" / \"스치니들 구매가격에 영향 전혀 없자라🩷\""],
 ["유도 장치","시선·클릭 유도","👇 ⬇️ \"보러가기\", \"반칙 재료\", \"비밀재료\""],
 ["인게이지먼트 베이트","알고리즘 부스팅","\"스+하+릿=초고속 2배반사\", \"스할완\", \"반하리완료\""],
 ["단축 링크","전환","link.coupang.com/a/XXXX (종종 2개 중복)"],
 ["해시태그","검색 유입","#선글라스 #파우치 #쿠팡파트너스 #네이버쇼핑"],
]
rr = hr + 1
for row in comp:
    for c, val in enumerate(row, 1):
        cell = ws3.cell(row=rr, column=c, value=val)
        cell.border = BORDER
        cell.alignment = WRAP
    rr += 1

# =========================================================
# Sheet 4: 수용도/메타반응
# =========================================================
ws4 = wb.create_sheet("수용도_메타반응")
ws4["A1"] = "커뮤니티 수용도 & 리스크 ('쿠파스' 검색 메타반응)"
ws4["A1"].font = TITLE_FONT
ws4["A2"] = "어그로 강도 ↔ 신뢰도 트레이드오프. 과하면 차단·저품질."
ws4["A2"].font = Font(italic=True, color="595959", size=10)
h4 = ["작성자", "발언", "시사점"]
HR4 = 4
for c, h in enumerate(h4, 1):
    ws4.cell(row=HR4, column=c, value=h)
style_header(ws4, HR4, len(h4))
data4 = [
 ["salt_peppermom_syd","\"비밀재료·비밀비법 뭐 이런거 좀 그만했음, 그냥 알려주면 안돼는지\" (109♥)","'비밀재료' 클리셰 피로 누적"],
 ["bn_love_00","\"다 오픈해서 알려준 사람 링크는 사는데, 안 알려주고 쿠파스만 걸면 차단\"","정보 게이팅 = 역효과"],
 ["danso.pohang","\"쿠파스 달고 올리는 컨텐츠 진짜 믿음 안 가네ㅋㅋ\"","신뢰도 저하"],
 ["daily_joodi","\"쿠파스 스레드로 하면 사람들이 싫어하고 아예 안 보거나 차단\"","도달 페널티 체감"],
 ["hz_labs","\"쿠파스 글 조회수 죽음. 너무 광고같음 → 더 실제같은 말투로 자동화 업데이트\"","'광고 티' 제거가 핵심 변수(자동화 운영자 자인)"],
 ["im_nerd_kim","\"에드센스 유입 어그로 … 쿠파스 어그로 같은거\"","어그로 통칭화"],
 ["dino_nogada","\"시간 없으면 스레드에만 집중, 벤치마킹하면 다 보임, 비싼 강의 필요없음\"","진입장벽 낮음 → 공급 과잉"],
]
r = HR4 + 1
for row in data4:
    for c, val in enumerate(row, 1):
        cell = ws4.cell(row=r, column=c, value=val)
        cell.border = BORDER
        cell.alignment = WRAP
    if (r - HR4) % 2 == 0:
        for c in range(1, len(h4) + 1):
            ws4.cell(row=r, column=c).fill = ALT_FILL
    r += 1
finalize(ws4, HR4, len(h4), [22, 60, 30])

wb.save(OUT)
print("SAVED", OUT)
